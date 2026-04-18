import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FOLDER = "outputs"
OUTPUT_FILE = "grading_sheet.xlsx"

COLUMNS = [
    "scenario_id",
    "scenario_title", 
    "target_model",
    "run_number",
    "turn_number",
    "teen_message",
    "assistant_message",
    # Human scores
    "h_dependence_framing",
    "h_boundary_setting",
    "h_resource_diversity",
    "h_secrecy_transparency",
    "h_notes",
    # LLM judge scores
    "llm_dependence_framing",
    "llm_boundary_setting",
    "llm_resource_diversity",
    "llm_secrecy_transparency",
    "llm_notes",
]

HEADER_LABELS = {
    "scenario_id": "Scenario ID",
    "scenario_title": "Scenario",
    "target_model": "Model",
    "run_number": "Run #",
    "turn_number": "Turn #",
    "teen_message": "Teen Message",
    "assistant_message": "Assistant Response",
    "h_dependence_framing": "H: Dependence Framing (0-2)",
    "h_boundary_setting": "H: Boundary Setting (0-2)",
    "h_resource_diversity": "H: Resource Diversity (0-2)",
    "h_secrecy_transparency": "H: Secrecy vs Transparency (0-2)",
    "h_notes": "H: Notes",
    "llm_dependence_framing": "LLM: Dependence Framing (0-2)",
    "llm_boundary_setting": "LLM: Boundary Setting (0-2)",
    "llm_resource_diversity": "LLM: Resource Diversity (0-2)",
    "llm_secrecy_transparency": "LLM: Secrecy vs Transparency (0-2)",
    "llm_notes": "LLM: Notes",
}

# Colors
HEADER_FILL = PatternFill("solid", start_color="1F4E79")  # dark blue
META_FILL = PatternFill("solid", start_color="D6E4F0")    # light blue - metadata cols
TEXT_FILL = PatternFill("solid", start_color="FFFFFF")     # white - message cols
HUMAN_FILL = PatternFill("solid", start_color="E2EFDA")   # light green - human scores
LLM_FILL = PatternFill("solid", start_color="FFF2CC")     # light yellow - llm scores
ALT_ROW = PatternFill("solid", start_color="F5F5F5")      # subtle gray for alternating

def load_conversations(folder):
    rows = []
    for filename in sorted(os.listdir(folder)):
        if not filename.endswith(".json"):
            continue
        filepath = os.path.join(folder, filename)
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        messages = data.get("messages", [])
        turn_number = 0
        last_teen_msg = ""

        for i, msg in enumerate(messages):
            if msg["role"] == "user":
                last_teen_msg = msg["content"]
            elif msg["role"] == "assistant":
                turn_number += 1
                rows.append({
                    "scenario_id": data.get("scenario_id", ""),
                    "scenario_title": data.get("scenario_title", ""),
                    "target_model": data.get("target_model", ""),
                    "run_number": data.get("run_number", ""),
                    "turn_number": turn_number,
                    "teen_message": last_teen_msg,
                    "assistant_message": msg["content"],
                    "h_dependence_framing": "",
                    "h_boundary_setting": "",
                    "h_resource_diversity": "",
                    "h_secrecy_transparency": "",
                    "h_notes": "",
                    "llm_dependence_framing": "",
                    "llm_boundary_setting": "",
                    "llm_resource_diversity": "",
                    "llm_secrecy_transparency": "",
                    "llm_notes": "",
                })
    return rows


def build_sheet(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grading"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = HEADER_LABELS[col_key]
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 40

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        is_alt = row_idx % 2 == 0
        for col_idx, col_key in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(col_key, "")
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            # Color coding by column group
            if col_key in ("scenario_id", "scenario_title", "target_model", "run_number", "turn_number"):
                cell.fill = ALT_ROW if is_alt else META_FILL
            elif col_key in ("teen_message", "assistant_message"):
                cell.fill = ALT_ROW if is_alt else PatternFill("solid", start_color="FAFAFA")
            elif col_key.startswith("h_"):
                cell.fill = HUMAN_FILL
            elif col_key.startswith("llm_"):
                cell.fill = LLM_FILL

        ws.row_dimensions[row_idx].height = 80

    # Column widths
    widths = {
        "scenario_id": 18,
        "scenario_title": 20,
        "target_model": 22,
        "run_number": 7,
        "turn_number": 7,
        "teen_message": 40,
        "assistant_message": 55,
        "h_dependence_framing": 12,
        "h_boundary_setting": 12,
        "h_resource_diversity": 12,
        "h_secrecy_transparency": 12,
        "h_notes": 30,
        "llm_dependence_framing": 12,
        "llm_boundary_setting": 12,
        "llm_resource_diversity": 12,
        "llm_secrecy_transparency": 12,
        "llm_notes": 30,
    }
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_key, 15)

    # Freeze top row and first 5 columns
    ws.freeze_panes = "F2"

    wb.save(OUTPUT_FILE)
    print(f"✓ Saved: {OUTPUT_FILE} ({len(rows)} rows)")


if __name__ == "__main__":
    rows = load_conversations(INPUT_FOLDER)
    if not rows:
        print("No JSON files found in outputs/ folder.")
    else:
        build_sheet(rows)